#!usr/bin/env python3
# _*_ coding: utf-8 _*_
"""
@Author: Joshua
@File: excel_reader.py
@Time: 2020/1/8
"""

import openpyxl
from compare2excels.utils import excel_logger
import sys
import os
from numpy import matrix


def parse_sheet_names(sheet_names):
    name1 = int(sheet_names[0].split("k")[0])
    name2 = int(sheet_names[1].split("k")[0])
    if name1 < name2:
        return str(sheet_names[0]), str(sheet_names[1])
    else:
        return str(sheet_names[1]), str(sheet_names[0])


def strip_excel(ws):
    max_row = ws.max_row
    max_column = ws.max_column
    row_num = 0
    column_num = 0
    if max_column < 1 or max_row < 1:
        excel_logger.warning("The work sheet {} is empty.")
    try:
        # 中间不允许有空行或者空列，去掉可能存在的空格行
        for row_num in range(1, max_row + 1):
            if row_num == 1:
                continue
            if ws.cell(row_num, 1).value == None:
                row_num -= 1
                break
            elif ws.cell(row_num, 1).value.strip() == "":
                row_num -= 1
                break
        for column_num in range(1, max_column + 1):
            if column_num == 1:
                continue
            if ws.cell(1, column_num).value == None:
                column_num -= 1
                break
            elif ws.cell(1, column_num).value.strip() == "":
                column_num -= 1
                break
    except IndexError as ie:
        excel_logger.error(ie.args)

    return row_num, column_num


def excel_to_dict(sheet_name, ws):
    row_num, column_num = strip_excel(ws)
    excel_logger.debug("Sheet {} has {} rows and {} columns.".format(sheet_name, row_num, column_num))
    all_nodes = []
    for i in range(2, column_num + 1):
        node = {}
        node_name = ws.cell(1, i).value.strip()
        node[node_name] = {}
        for j in range(2, row_num + 1):
            attr_key = str(ws.cell(j, 1).value).strip()
            attr_value = str(ws.cell(j, i).value).strip().split(";")
            node[node_name].update({attr_key: attr_value})
        all_nodes.append(node)
    return all_nodes


def read_excels(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    sheet_names = workbook.sheetnames  # .get_sheet_names()
    source, target = parse_sheet_names(sheet_names)
    excel_logger.debug("Source is {}, target is {}.".format(source, target))
    source_dict = excel_to_dict(source, workbook[source])
    target_dict = excel_to_dict(target, workbook[target])
    excel_logger.debug(source_dict)
    excel_logger.debug(target_dict)
    return source_dict, target_dict


